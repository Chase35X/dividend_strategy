import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule


def getDividendData(ticker):
    stock = yf.Ticker(ticker)
    dividends = stock.get_dividends()

    dividends_df = pd.DataFrame(dividends)

    dividends_df['Year'] = dividends.index.year
    dividends_df = dividends_df[['Dividends', 'Year']]

    dividend_list = dividends_df.values.tolist()

    dividendDictionary = {}


    for dividend in dividend_list:
        if dividend[1] in dividendDictionary:
            dividendDictionary[str(int(dividend[1]))] += float(dividend[0])
        else:
            dividendDictionary[str(int(dividend[1]))] = float(dividend[0])
    
    return dividendDictionary

def getPriceData(ticker):
    data = yf.download(ticker, '1987-01-01', '2023-12-31', interval="1mo")

    price_df = pd.DataFrame(data)
    price_df['Year'] = price_df.index.year
    price_df = price_df[['High', 'Year']]

    price_list = price_df.values.tolist()

    priceDictionary = {}

    for price in price_list:
        if str(int(price[1])) in priceDictionary:
            priceDictionary[str(int(price[1]))] += float(price[0])
        else:
            priceDictionary[str(int(price[1]))] = float(price[0])
    
    for price in priceDictionary:
        priceDictionary[price] = priceDictionary[price] / 12
    
    return priceDictionary

def getStockSplitData(ticker):
    stock = yf.Ticker(ticker)
    stockSplits = stock.get_splits()

    splits_df = pd.DataFrame(stockSplits)

    splits_df['Year'] = stockSplits.index.year
    splits_df = splits_df[['Stock Splits', 'Year']]

    splits_list = splits_df.values.tolist()

    splitsDictionary = {}

    for split in splits_list:
        if str(int(split[1])) in splitsDictionary:
            splitsDictionary[str(int(split[1]))] += ', ' + str(int(split[0])) + ' for 1'
        else:
            splitsDictionary[str(int(split[1]))] = str(int(split[0])) + ' for 1'

    return splitsDictionary


    
    


def getStockInformation(ticker):
    stock = yf.Ticker(ticker)

    informationDictionary = {}
    informationDictionary['name'] = stock.info['shortName']
    informationDictionary['stock_price'] = stock.info['currentPrice']
    informationDictionary['free_Cashflow'] = stock.info['freeCashflow']

    return informationDictionary

def getIndividualDividend(dividendData, year):
    if year in dividendData:
        return dividendData[year]
    else:
        return ''

def getIndividualPrice(priceData, year):
    if year in priceData:
        return priceData[year]
    else:
        return ''

def getIndividualYield(priceData, dividendData, year):
    if year in dividendData and year in priceData:
        return dividendData[year] / priceData[year]
    else:
        return ''

def getIndividualSplit(splitData, year):
    if year in splitData:
        return splitData[year]
    else:
        return ''

def getPercentChange(dividendData, year):
    if year in dividendData and str(int(year) - 1) in dividendData:
        return (dividendData[year] - dividendData[str(int(year) - 1)]) / dividendData[str(int(year) - 1)]
    else:
        return ''

def getStockAllData(ticker, ws):

    informationData = getStockInformation(ticker)
    dividendData = getDividendData(ticker)
    priceData = getPriceData(ticker)
    splitData = getStockSplitData(ticker)

    # Labels
    ws.append(['Ticker','Stock','Current Price',1987,1988,1989,1990,1991,1992,1993,1994,1995,1996,1997,1998,1999,2000,2001,2002,2003,2004,2005,2006,2007,2008,2009,2010,2011,2012,2013,2014,2015,2016,2017,2018,2019,2020,2021,2022,2023,2024,'','','Current Year Dividend Analysis','','','Forecasts'])
    
    # Basic information and dividends
    ws.append([ticker, informationData['name'], informationData['stock_price'], getIndividualDividend(dividendData,'1987'), getIndividualDividend(dividendData,'1988'), getIndividualDividend(dividendData,'1989'), getIndividualDividend(dividendData,'1990'), getIndividualDividend(dividendData,'1991'), getIndividualDividend(dividendData,'1992'), getIndividualDividend(dividendData,'1993'), getIndividualDividend(dividendData,'1994'), getIndividualDividend(dividendData,'1995'), getIndividualDividend(dividendData,'1996'), getIndividualDividend(dividendData,'1997'), getIndividualDividend(dividendData,'1998'), getIndividualDividend(dividendData,'1999'), getIndividualDividend(dividendData,'2000'), getIndividualDividend(dividendData,'2001'), getIndividualDividend(dividendData,'2002'), getIndividualDividend(dividendData,'2003'), getIndividualDividend(dividendData,'2004'), getIndividualDividend(dividendData,'2005'), getIndividualDividend(dividendData,'2006'), getIndividualDividend(dividendData,'2007'), getIndividualDividend(dividendData,'2008'), getIndividualDividend(dividendData,'2009'), getIndividualDividend(dividendData,'2010'), getIndividualDividend(dividendData,'2011'), getIndividualDividend(dividendData,'2012'), getIndividualDividend(dividendData,'2013'), getIndividualDividend(dividendData,'2014'), getIndividualDividend(dividendData,'2015'), getIndividualDividend(dividendData,'2016'), getIndividualDividend(dividendData,'2017'), getIndividualDividend(dividendData,'2018'), getIndividualDividend(dividendData,'2019'), getIndividualDividend(dividendData,'2020'), getIndividualDividend(dividendData,'2021'), getIndividualDividend(dividendData,'2022'), getIndividualDividend(dividendData,'2023'), getIndividualDividend(dividendData,'2024'),'','','Q1'])
    
    # Percent Changes
    ws.append(['','','% Change','', getPercentChange(dividendData,'1988'),getPercentChange(dividendData,'1989'),getPercentChange(dividendData,'1990'),getPercentChange(dividendData,'1991'),getPercentChange(dividendData,'1992'),getPercentChange(dividendData,'1993'),getPercentChange(dividendData,'1994'),getPercentChange(dividendData,'1995'),getPercentChange(dividendData,'1996'),getPercentChange(dividendData,'1997'),getPercentChange(dividendData,'1998'),getPercentChange(dividendData,'1999'),getPercentChange(dividendData,'2000'),getPercentChange(dividendData,'2001'),getPercentChange(dividendData,'2002'),getPercentChange(dividendData,'2003'),getPercentChange(dividendData,'2004'),getPercentChange(dividendData,'2005'),getPercentChange(dividendData,'2006'),getPercentChange(dividendData,'2007'),getPercentChange(dividendData,'2008'),getPercentChange(dividendData,'2009'),getPercentChange(dividendData,'2010'),getPercentChange(dividendData,'2011'),getPercentChange(dividendData,'2012'),getPercentChange(dividendData,'2013'),getPercentChange(dividendData,'2014'),getPercentChange(dividendData,'2015'),getPercentChange(dividendData,'2016'),getPercentChange(dividendData,'2017'),getPercentChange(dividendData,'2018'),getPercentChange(dividendData,'2019'),getPercentChange(dividendData,'2020'),getPercentChange(dividendData,'2021'),getPercentChange(dividendData,'2022'),getPercentChange(dividendData,'2023'),'','','', 'Q2'])

    # Prices
    ws.append(['','','Average Price', getIndividualPrice(priceData,'1987'), getIndividualPrice(priceData,'1988'), getIndividualPrice(priceData,'1989'), getIndividualPrice(priceData,'1990'), getIndividualPrice(priceData,'1991'), getIndividualPrice(priceData,'1992'), getIndividualPrice(priceData,'1993'), getIndividualPrice(priceData,'1994'), getIndividualPrice(priceData,'1995'), getIndividualPrice(priceData,'1996'), getIndividualPrice(priceData,'1997'), getIndividualPrice(priceData,'1998'), getIndividualPrice(priceData,'1999'), getIndividualPrice(priceData,'2000'), getIndividualPrice(priceData,'2001'), getIndividualPrice(priceData,'2002'), getIndividualPrice(priceData,'2003'), getIndividualPrice(priceData,'2004'), getIndividualPrice(priceData,'2005'), getIndividualPrice(priceData,'2006'), getIndividualPrice(priceData,'2007'), getIndividualPrice(priceData,'2008'), getIndividualPrice(priceData,'2009'), getIndividualPrice(priceData,'2010'), getIndividualPrice(priceData,'2011'), getIndividualPrice(priceData,'2012'), getIndividualPrice(priceData,'2013'), getIndividualPrice(priceData,'2014'), getIndividualPrice(priceData,'2015'), getIndividualPrice(priceData,'2016'), getIndividualPrice(priceData,'2017'), getIndividualPrice(priceData,'2018'), getIndividualPrice(priceData,'2019'), getIndividualPrice(priceData,'2020'), getIndividualPrice(priceData,'2021'), getIndividualPrice(priceData,'2022'), getIndividualPrice(priceData,'2023'), '','','','Q3'])
    
    # Dividend Yield
    ws.append(['','','Dividend Yield', getIndividualYield(priceData, dividendData, '1987'),getIndividualYield(priceData, dividendData, '1988'),getIndividualYield(priceData, dividendData, '1989'),getIndividualYield(priceData, dividendData, '1990'),getIndividualYield(priceData, dividendData, '1991'),getIndividualYield(priceData, dividendData, '1992'),getIndividualYield(priceData, dividendData, '1993'),getIndividualYield(priceData, dividendData, '1994'),getIndividualYield(priceData, dividendData, '1995'),getIndividualYield(priceData, dividendData, '1996'),getIndividualYield(priceData, dividendData, '1997'),getIndividualYield(priceData, dividendData, '1998'),getIndividualYield(priceData, dividendData, '1999'),getIndividualYield(priceData, dividendData, '2000'),getIndividualYield(priceData, dividendData, '2001'),getIndividualYield(priceData, dividendData, '2002'),getIndividualYield(priceData, dividendData, '2003'),getIndividualYield(priceData, dividendData, '2004'),getIndividualYield(priceData, dividendData, '2005'),getIndividualYield(priceData, dividendData, '2006'),getIndividualYield(priceData, dividendData, '2007'),getIndividualYield(priceData, dividendData, '2008'),getIndividualYield(priceData, dividendData, '2009'),getIndividualYield(priceData, dividendData, '2010'),getIndividualYield(priceData, dividendData, '2011'),getIndividualYield(priceData, dividendData, '2012'),getIndividualYield(priceData, dividendData, '2013'),getIndividualYield(priceData, dividendData, '2014'),getIndividualYield(priceData, dividendData, '2015'),getIndividualYield(priceData, dividendData, '2016'),getIndividualYield(priceData, dividendData, '2017'),getIndividualYield(priceData, dividendData, '2018'),getIndividualYield(priceData, dividendData, '2019'),getIndividualYield(priceData, dividendData, '2020'),getIndividualYield(priceData, dividendData, '2021'),getIndividualYield(priceData, dividendData, '2022'),getIndividualYield(priceData, dividendData, '2023'),'','','','Q4'])

    # Stock Splits
    ws.append(['','', 'Stock Splits', getIndividualSplit(splitData, '1987'), getIndividualSplit(splitData, '1988'), getIndividualSplit(splitData, '1989'), getIndividualSplit(splitData, '1990'), getIndividualSplit(splitData, '1991'), getIndividualSplit(splitData, '1992'), getIndividualSplit(splitData, '1993'), getIndividualSplit(splitData, '1994'), getIndividualSplit(splitData, '1995'), getIndividualSplit(splitData, '1996'), getIndividualSplit(splitData, '1997'), getIndividualSplit(splitData, '1998'), getIndividualSplit(splitData, '1999'), getIndividualSplit(splitData, '2000'), getIndividualSplit(splitData, '2001'), getIndividualSplit(splitData, '2002'), getIndividualSplit(splitData, '2003'), getIndividualSplit(splitData, '2004'), getIndividualSplit(splitData, '2005'), getIndividualSplit(splitData, '2006'), getIndividualSplit(splitData, '2007'), getIndividualSplit(splitData, '2008'), getIndividualSplit(splitData, '2009'), getIndividualSplit(splitData, '2010'), getIndividualSplit(splitData, '2011'), getIndividualSplit(splitData, '2012'), getIndividualSplit(splitData, '2013'), getIndividualSplit(splitData, '2014'), getIndividualSplit(splitData, '2015'), getIndividualSplit(splitData, '2016'), getIndividualSplit(splitData, '2017'), getIndividualSplit(splitData, '2018'), getIndividualSplit(splitData, '2019'), getIndividualSplit(splitData, '2020'), getIndividualSplit(splitData, '2021'), getIndividualSplit(splitData, '2022'), getIndividualSplit(splitData, '2023'), getIndividualSplit(splitData, '2024'), '', '', 'Dividends So Far'])

    # Blank line
    ws.append([''])


def edit_sheet(ws):
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['AR'].width = 35
    thin_black_border = Side(border_style='thin', color ='000000')

    # Format labels as bold
    for i in range(1, 1000, 7):
        for cell in ws[str(i) + ":" + str(i)]:
            cell.font = Font(bold=True, size = 12)
    
    # Format vertical labels as bold and right side border
    for k in range(3,1000,7):
        for c in range(k,k+4):
            ws.cell(row=c, column=3).font = Font(bold=True, size = 12)
            ws.cell(row=c, column=3).border = Border(right=thin_black_border)

    # format labels as having a bottom border
    for j in range(1, 1000, 7):
        for cell in ws[str(j) + ":" + str(j)]:
            cell.border = Border(bottom=thin_black_border)

    # Format labels as aligned center
    for m in range(1,1000):
        ws.cell(row=m, column=3).alignment = Alignment(horizontal='center')
    

    greenFill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    redFill = PatternFill(start_color='FF474C', end_color='FF474C', fill_type='solid')
    

    # Format changes in dividends as percentages and conditional formatting for percent changes
    for p in range(3,1000,7):
        for c in range(4,41):
            ws.cell(row=p, column=c).number_format = '0.00%'
            ws.conditional_formatting.add('E' + str(p) + ':AN' + str(p), CellIsRule(operator='greaterThan', formula=[0], fill=greenFill))
            ws.conditional_formatting.add('E' + str(p) + ':AN' + str(p), CellIsRule(operator='lessThan', formula=[0], fill=redFill))

    # Format dividend yield as a percentage
    for y in range(5,1000,7):
        for c in range(4,41):
            ws.cell(row=y, column=c).number_format = '0.00%'

    # Format average price as a $
    for price in range(4,1000,7):
        for c in range(4,41):
            ws.cell(row=price, column=c).number_format = '$0.00'

    # Format current price as a $
    for cprice in range(2, 1000, 7):
        ws.cell(row=cprice, column=3).number_format = '$0.00'








def make_sheet(tickerList):
    wb = Workbook()

    # This will need to be changed to the local persons folder
    # filepath = '/Users/chase/Documents'

    # wb.save(filepath)
    ws = wb.active

    for ticker in tickerList:
        getStockAllData(ticker, ws)

    edit_sheet(ws)

    wb.save("sample.xlsx")
    return wb


# tickerList = ["MSFT","NVDA","AAPL","AMZN","META","GOOGL","GOOG","BRK","B","LLY","JPM","AVGO","XOM","TSLA","UNH","V","PG","MA","COST","JNJ","HD","MRK","ABBV","WMT","BAC","NFLX","CVX","AMD","KO","PEP","QCOM","CRM","TMO","WFC","LIN","ADBE","ORCL","CSCO","MCD","DIS","ABT","AMAT","ACN","TXN","GE","VZ","DHR","CAT","PFE","AMGN","PM","NEE","INTU","CMCSA","IBM","GS","RTX","ISRG","MU","UNP","SPGI","AXP","NOW","COP","HON","BKNG","UBER","ETN","T","INTC","ELV","LOW","LRCX","PGR","MS","VRTX","TJX","C","NKE","SYK","ADI","BSX","MDT","BLK","CB","SCHW","BA","REGN","KLAC","MMC","LMT","ADP","UPS","CI","PLD","DE","SBUX","PANW","AMT","MDLZ","TMUS","FI","SO","BX","SNPS","BMY","CMG","MO","DUK","ZTS","GILD","APH","CDNS","ICE","CL","CVS","MCK","FCX","ANET","TDG","WM","TT","CME","SHW","TGT","EQIX","EOG","NXPI","BDX","PYPL","GD","CEG","PH","HCA","CSX","ITW","MPC","ABNB","NOC","MCO","SLB","EMR","USB","PNC","MSI","APD","PSX","ECL","CTAS","WELL","FDX","ROP","ORLY","MAR","AON","PCAR","MMM","AIG","AJG","EW","GM","VLO","COF","CARR","MCHP","NSC","HLT","WMB","SPG","TFC","MRNA","JCI","SRE","NEM","TRV","AZO","ROST","F","AEP","AFL","OKE","GEV","TEL","DLR","CPRT","KMB","BK","FIS","ADSK","D","CCI","HUM","DXCM","O","DHI","MET","PSA","AMP","PRU","ALL","URI","LHX","HES","NUE","IDXX","STZ","OTIS","OXY","LEN","IQV","PWR","DOW","GWW","YUM","CTVA","PCG","MSCI","SMCI","PAYX","GIS","A","AME","COR","MNST","CNC","RSG","ACGL","KMI","CMI","FTNT","PEG","EXC","KVUE","VRSK","FAST","IR","SYY","KDP","RCL","LULU","MPWR","MLM","DD","FANG","KR","VMC","BIIB","XYL","HWM","ADM","IT","CTSH","GEHC","DAL","EA","ED","BKR","FICO","CSGP","ON","VST","HAL","PPG","DFS","HPQ","EXR","DG","HIG","RMD","XEL","ODFL","MTD","DVN","CDW","VICI","WAB","ROK","HSY","EIX","FSLR","TSCO","EL","GLW","EFX","CHTR","KHC","DECK","EBAY","ANSS","AVB","WTW","CHD","TROW","TRGP","TTWO","GPN","CBRE","FTV","WEC","DOV","AWK","DLTR","FITB","GRMN","NTAP","MTB","IFF","PHM","CAH","WST","NVR","LYB","WDC","DTE","KEYS","ZBH","ETR","APTV","IRM","BR","HPE","RJF","STT","STE","EQR","NDAQ","BALL","VLTO","WY","TER","PPL","SBAC","BRO","ES","CTRA","FE","PTC","HUBB","GPC","STLD","VTR","INVH","LDOS","TYL","AXON","HBAN","CNP","AEE","ULTA","BLDR","COO","CPAY","TDY","WBD","CBOE","ARE","WAT","AVY","CMS","CINF","DPZ","DRI","ALGN","MKC","MOH","SYF","OMC","PFG","NRG","STX","EXPD","RF","HOLX","J","ENPH","NTRS","UAL","BAX","ATO","TXT","BBY","EQT","ESS","EG","MRO","LVS","LH","LUV","PKG","WRB","ILMN","TSN","CLX","CFG","CCL","K","ZBRA","IP","DGX","BG","VRSN","IEX","MAA","CF","MAS","EXPE","FDS","JBL","AMCR","CE","SWKS","SNA","ALB","CAG","DOC","GEN","POOL","AES","WRK","L","AKAM","TRMB","RVTY","LYV","SWK","KEY","PNR","JBHT","KIM","ROL","LNT","HST","VTRS","PODD","LW","EVRG","NDSN","TECH","JKHY","BBWI","UDR","IPG","EMN","NI","LKQ","WBA","SJM","UHS","JNPR","KMX","CPT","MGM","INCY","CRL","ALLE","NWSA","REG","CHRW","TPR","TFX","AOS","EPAM","MOS","HII","TAP","FFIV","CTLT","QRVO","HSIC","WYNN","AIZ","APA","HRL","CPB","GNRC","PNW","FOXA","BXP","BWA","MTCH","BF","B","SOLV","DVA","ETSY","DAY","CZR","AAL","HAS","MKTX","FRT","FMC","RL","NCLH","PAYC","GL","IVZ","RHI","BEN","CMA","MHK","PARA","BIO","FOX","NWS"]
# make_sheet(tickerList)

# from flask import Flask, request, jsonify

# app = Flask(__name__)

@app.route('/make_sheet', methods=['GET'])
def call_my_function():
    ticker_list = request.args.get('param', type=list)
    result = make_sheet(ticker_list)
    return result

# if __name__ == '__main__':
#     app.run()

